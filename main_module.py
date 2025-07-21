
import os
import sys
import json
import requests
import re
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import matplotlib.pyplot as plt

def get_app_path():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS  # 正確使用 PyInstaller 的路徑
    return os.path.dirname(os.path.abspath(__file__))

app_dir = get_app_path()
config_path = os.path.join(app_dir, "config.json")

app_dir = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(app_dir, "config.json")
with open(config_path, "r", encoding="utf-8") as f:
    config = json.load(f)

START_YEAR = config.get("start_year", 2024)
END_YEAR = config.get("end_year", 2025)
MONTHS = config.get("months", list(range(1, 13)))
ENABLE_STATS = config.get("enable_stats", True)
ENABLE_CHART = config.get("enable_chart", False)
EXCEL_FILE = os.path.join(app_dir, "539_by_year.xlsx")
TRANSITION_FILE = os.path.join(app_dir, "539_transition_analysis.txt")
CHART_FILE = os.path.join(app_dir, "539_multiples_of_3_chart.png")

def fetch_data(year, month):
    month_str = f"{year}-{month:02d}"
    url = f"https://api.taiwanlottery.com/TLCAPIWeB/Lottery/Daily539Result?period&month={month_str}&pageNum=1&pageSize=50"
    try:
        res = requests.get(url, verify=False)
        res.raise_for_status()
        return res.json()['content']['daily539Res']
    except:
        return []

def fetch_today_data():
    today = datetime.today()
    month_str = today.strftime("%Y-%m")
    date_str = today.strftime("%Y-%m-%d")
    url = f"https://api.taiwanlottery.com/TLCAPIWeB/Lottery/Daily539Result?period&month={month_str}&pageNum=1&pageSize=50"
    try:
        res = requests.get(url, verify=False)
        res.raise_for_status()
        for r in res.json()['content']['daily539Res']:
            if r['lotteryDate'].startswith(date_str):
                return r
    except:
        return None

def prepare_workbook():
    if os.path.exists(EXCEL_FILE):
        return load_workbook(EXCEL_FILE)
    wb = Workbook()
    wb.remove(wb.active)
    return wb

def get_existing_dates(ws):
    return set(str(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row[0])

def save_to_excel(records_by_year):
    wb = prepare_workbook()
    for year, records in records_by_year.items():
        sheet_name = str(year)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["開獎日", "號碼1", "號碼2", "號碼3", "號碼4", "號碼5"])
        existing_dates = get_existing_dates(ws)
        sorted_records = sorted(records, key=lambda r: r['lotteryDate'])
        for r in sorted_records:
            date = r['lotteryDate'].split("T")[0]
            if date not in existing_dates:
                ws.append([date] + r['drawNumberSize'])
    wb.save(EXCEL_FILE)

def update_history():
    records_by_year = {}
    for year in range(START_YEAR, END_YEAR + 1):
        for month in MONTHS:
            records = fetch_data(year, month)
            if records:
                records_by_year.setdefault(year, []).extend(records)
    save_to_excel(records_by_year)

def update_today():
    record = fetch_today_data()
    if record:
        year = datetime.today().year
        save_to_excel({year: [record]})
        return True
    return False

def is_multiple_of_3(n):
    return n % 3 == 0

def generate_stats():
    wb = load_workbook(EXCEL_FILE)
    for name in wb.sheetnames[:]:
        if name.endswith("統計"):
            del wb[name]
    for name in wb.sheetnames:
        if name.isdigit():
            ws = wb[name]
            counter = Counter()
            for row in ws.iter_rows(min_row=2, values_only=True):
                counter.update(row[1:6])
            stat_ws = wb.create_sheet(title=name + "統計")
            stat_ws.append(["號碼", "出現次數"])
            for num in sorted(counter):
                stat_ws.append([num, counter[num]])
                if is_multiple_of_3(num):
                    stat_ws[f"A{stat_ws.max_row}"].font = Font(color="FF0000")
    wb.save(EXCEL_FILE)

def generate_multiples_of_3_chart():
    wb = load_workbook(EXCEL_FILE)
    counter = Counter()
    for sheet in wb.sheetnames:
        if sheet.isdigit():
            for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                counter.update(row[1:6])
    data = {n: counter[n] for n in range(3, 40, 3)}
    nums, counts = list(data.keys()), list(data.values())
    plt.figure(figsize=(10, 5))
    bars = plt.bar([str(n) for n in nums], counts, color='red')
    plt.title("今彩539 - 3 的倍數號碼出現次數")
    plt.xlabel("號碼")
    plt.ylabel("出現次數")
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2, height + 0.5, str(int(height)),
                 ha='center', va='bottom', fontsize=9)
    plt.tight_layout()
    plt.savefig(CHART_FILE)
    plt.show()

def analyze_transition_patterns():
    wb = load_workbook(EXCEL_FILE)
    records = []
    for sheet_name in sorted(wb.sheetnames):
        if sheet_name.isdigit():
            for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
                if all(isinstance(n, int) for n in row[1:6]):
                    records.append((row[0], set(row[1:6])))
    transitions = defaultdict(Counter)
    for i in range(len(records) - 1):
        for num in records[i][1]:
            transitions[num].update(records[i + 1][1])
    with open(TRANSITION_FILE, "w", encoding="utf-8") as f:
        for num in range(1, 40):
            if num in transitions:
                f.write(f"🔁 當期號碼 {num} 出現時，下一期常見號碼：\n")
                for follow_num, count in transitions[num].most_common(10):
                    f.write(f"    - {follow_num:02d}：出現 {count} 次\n")
                f.write("\n")


def recommend_by_transition():
    if not os.path.exists(TRANSITION_FILE):
        return None
    transitions = {}
    with open(TRANSITION_FILE, "r", encoding="utf-8") as f:
        current_key = None
        for line in f:
            if line.startswith("🔁"):
                match = re.search(r"號碼\s*(\d+)", line)
                if not match:
                    match = re.search(r"(\d+)", line)
                if match:
                    current_key = int(match.group(1))
                    transitions[current_key] = []
            elif line.strip().startswith("-") and current_key is not None:
                parts = line.strip().split("：")
                num = int(parts[0].split("-")[1])
                count = int(parts[1].replace("出現", "").replace("次", "").strip())
                transitions[current_key].append((num, count))

    wb = load_workbook(EXCEL_FILE)
    sheets = sorted([s for s in wb.sheetnames if s.isdigit()])
    latest_sheet = wb[sheets[-1]]
    rows = list(latest_sheet.iter_rows(min_row=2, values_only=True))
    last_row = rows[-1]
    last_nums = last_row[1:6]
    counter = Counter()
    for n in last_nums:
        for to_num, score in transitions.get(n, []):
            counter[to_num] += score
    for n in last_nums:
        counter.pop(n, None)
    top10_all = counter.most_common(10)
    top10 = [num for num, _ in top10_all]
    top5 = [num for num, _ in top10_all[:5]]
    return last_nums, sorted(top10), top5


